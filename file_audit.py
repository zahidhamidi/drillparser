import os
from openpyxl import load_workbook
from PyPDF2 import PdfReader
from xls2xlsx import XLS2XLSX

import xlrd
import pandas as pd
from collections import defaultdict
import streamlit as st



def is_file_corrupted(directory_path):
    corrupted_files_count = 0
    corrupted_files_info = {}

    for file_name in os.listdir(directory_path):
        file_path = os.path.join(directory_path, file_name)

        

        try:
            # Try to read the file using the appropriate library (e.g., pandas for CSV files)
            if file_name.endswith('.csv'):
                pd.read_csv(file_path)
            elif file_name.endswith('.txt'):
                # Add code to handle text file processing
                pass
            # Add more conditions for other file types as needed
        except Exception as e:
            # If an exception is raised, consider the file corrupted
            st.write(f"Error processing file {file_path}: {e}")

            # Increment the count and store file information
            corrupted_files_count += 1
            corrupted_files_info[file_name] = file_path

    return corrupted_files_count, corrupted_files_info


def find_duplicate_files(directory):
    file_info_dict = defaultdict(list)

    # Iterate through files in the directory
    for file_name in os.listdir(directory):
        file_path = os.path.join(directory, file_name)

        # Extract relevant file information
        file_size = os.path.getsize(file_path)
        file_type = file_name.split('.')[-1].lower()

        # Store file information in a dictionary
        file_info_dict[(file_name, file_type, file_size)].append(file_path)

        

    # Identify duplicates
    duplicates = {key: value for key, value in file_info_dict.items() if len(value) > 1}

    return duplicates

def count_pdf_pages(file_path):
    with open(file_path, 'rb') as file:
        pdf_reader = PdfReader(file)
        return len(pdf_reader.pages)

def analyze_pdf_files(temp_directory):
    pdf_files = [file for file in os.listdir(temp_directory) if file.endswith(".pdf")]

    pdf_file_stats = {}
    for pdf_file in pdf_files:
        file_path = os.path.join(temp_directory, pdf_file)
        num_pages = count_pdf_pages(file_path)
        file_size = os.path.getsize(file_path) / (1024 * 1024)  # Convert to MB
        pdf_file_stats[pdf_file] = {'Num Pages': num_pages, 'File Size (MB)': file_size}

    return pdf_file_stats

def count_files_and_sizes(directory):
    file_types = {'excel_xlsx': 0, 'excel_xls': 0, 'pdf': 0, 'word': 0, 'others': 0}
    file_sizes = {'excel_xlsx': 0, 'excel_xls': 0, 'pdf': 0, 'word': 0, 'others': 0}
    files_info = []  # List to store file information

    for file_name in os.listdir(directory):
        file_path = os.path.join(directory, file_name)
        if os.path.isfile(file_path):
            file_extension = file_name.split('.')[-1].lower()
            file_size = os.path.getsize(file_path) / (1024 * 1024)  # Size in MB

            # Update file types and sizes
            if file_extension == 'xlsx':
                file_types['excel_xlsx'] += 1
                file_sizes['excel_xlsx'] += file_size
            elif file_extension == 'xls':
                file_types['excel_xls'] += 1
                file_sizes['excel_xls'] += file_size
            elif file_extension == 'pdf':
                file_types['pdf'] += 1
                file_sizes['pdf'] += file_size
            elif file_extension in ['doc', 'docx']:
                file_types['word'] += 1
                file_sizes['word'] += file_size
            else:
                file_types['others'] += 1
                file_sizes['others'] += file_size

            # Append file information to the list
            files_info.append({'File Name': file_name, 'Size (MB)': file_size, 'Type': file_extension})

    return file_types, file_sizes, files_info


def convert_xls_to_xlsx(directory):
    for file_name in os.listdir(directory):
        if file_name.endswith('.xls'):
            x2x = XLS2XLSX(file_name)
            x2x.to_xlsx("f{file_name}.xlsx")
            file_path = os.path.join(directory, file_name)
            # output_path = os.path.join(directory, f'{os.path.splitext(file_name)[0]}.xlsx')

            # # Read the .xls file using pandas and save as .xlsx
            # df = pd.read_excel(file_path, engine='xlrd')
            # df.to_excel(output_path, index=False, engine='openpyxl')

            # Remove the original .xls file
            os.remove(file_path)


def analyze_excel_sheets(directory):
    total_sheets_xlsx = 0
    max_sheets_xlsx = 0
    min_sheets_xlsx = float('inf')
    max_sheets_xlsx_file = ""
    min_sheets_xlsx_file = ""

    total_sheets_xls = 0
    max_sheets_xls = 0
    min_sheets_xls = float('inf')
    max_sheets_xls_file = ""
    min_sheets_xls_file = ""

    sheet_count_list = []
    data = []  # To store data for the table

    for file_name in os.listdir(directory):
        if file_name.endswith('.xlsx'):
            file_path = os.path.join(directory, file_name)

            workbook = load_workbook(file_path, read_only=True)
            sheet_count = len(workbook.sheetnames)

            total_sheets_xlsx += sheet_count
            if sheet_count > max_sheets_xlsx:
                max_sheets_xlsx = sheet_count
                max_sheets_xlsx_file = file_name
            if sheet_count < min_sheets_xlsx:
                min_sheets_xlsx = sheet_count
                min_sheets_xlsx_file = file_name
            
            sheet_count_list.append(sheet_count)

            # Add data to the table
            data.append({
                'File Name': file_name,
                'File Size (MB)': os.path.getsize(file_path) / (1024 * 1024),  # Convert bytes to megabytes
                'File Type': 'xls',
                'Sheet Count': sheet_count
})

            

        elif file_name.endswith('.xls'):
            file_path = os.path.join(directory, file_name)

            workbook = xlrd.open_workbook(file_path)
            sheet_count = workbook.nsheets

            total_sheets_xls += sheet_count
            if sheet_count > max_sheets_xls:
                max_sheets_xls = sheet_count
                max_sheets_xls_file = file_name
            if sheet_count < min_sheets_xls:
                min_sheets_xls = sheet_count
                min_sheets_xls_file = file_name
            
            sheet_count_list.append(sheet_count)


            # Add data to the table
            data.append({
                'File Name': file_name,
                'File Size (MB)': os.path.getsize(file_path) / (1024 * 1024),  # Convert bytes to megabytes
                'File Type': 'xls',
                'Sheet Count': sheet_count
})



    if total_sheets_xlsx == 0:
        average_sheets_xlsx = 0
    else:
        average_sheets_xlsx = total_sheets_xlsx / len([file_name for file_name in os.listdir(directory) if file_name.endswith('.xlsx')])



    if total_sheets_xls == 0:
        average_sheets_xls = 0
    else:
        average_sheets_xls = total_sheets_xls / len([file_name for file_name in os.listdir(directory) if file_name.endswith('.xls')])

    return total_sheets_xlsx, max_sheets_xlsx, min_sheets_xlsx, average_sheets_xlsx, total_sheets_xls, max_sheets_xls, min_sheets_xls, average_sheets_xls, max_sheets_xlsx_file, min_sheets_xlsx_file, max_sheets_xls_file, min_sheets_xls_file,data
    

# Example usage
directory_path = r'C:\Users\ZHamid2\Downloads\Halini Files\Halini Deep-1 DDR'

"""
# Count files and sizes
file_types, file_sizes ,file_name = count_files_and_sizes(directory_path)
print("\n-------------------File Statistics-------------------")
print(f"File Types: {file_types}")
print(f"File Sizes (MB): {file_sizes}")

# Convert .xls to .xlsx
convert_xls_to_xlsx(directory_path)

# Analyze Excel sheets
(
    total_sheets_xlsx, max_sheets_xlsx, min_sheets_xlsx, average_sheets_xlsx,
    total_sheets_xls, max_sheets_xls, min_sheets_xls, average_sheets_xls,
    max_sheets_xlsx_file, min_sheets_xlsx_file, max_sheets_xls_file, min_sheets_xls_file
) = analyze_excel_sheets(directory_path)

print("\n-------------------Excel Sheet-------------------")
print(f"Total Sheets (.xlsx): {total_sheets_xlsx}")
print(f"Max Sheets (.xlsx): {max_sheets_xlsx} (File: {max_sheets_xlsx_file})")
print(f"Min Sheets (.xlsx): {min_sheets_xlsx} (File: {min_sheets_xlsx_file})")
print(f"Average Sheets (.xlsx): {average_sheets_xlsx}")

print("\n-------------------Old Excel Sheet-------------------")
print(f"Total Sheets (.xls): {total_sheets_xls}")
print(f"Max Sheets (.xls): {max_sheets_xls} (File: {max_sheets_xls_file})")
print(f"Min Sheets (.xls): {min_sheets_xls} (File: {min_sheets_xls_file})")
print(f"Average Sheets (.xls): {average_sheets_xls}")
"""